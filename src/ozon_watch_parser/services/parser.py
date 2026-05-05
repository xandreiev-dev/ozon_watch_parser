from __future__ import annotations

import asyncio
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from ozon_watch_parser.browser import BrowserManager, BrowserSettings
from ozon_watch_parser.config import BRAND_MIN_CARDS, BRAND_URLS, DEFAULT_MIN_CARDS
from ozon_watch_parser.domain import normalize_listing_item
from ozon_watch_parser.export import StreamingXlsxWriter, aggregate_brand_exports, brand_file_name
from ozon_watch_parser.ozon import ListingExtractor, ListingItem
from ozon_watch_parser.utils.dates import next_run_at
from ozon_watch_parser.utils.url import article_from_url, build_page_url, listing_url_variants

logger = logging.getLogger(__name__)


@dataclass(slots=True)
class ParseJob:
    brand: str
    source_urls: list[str]


class OzonWatchParser:
    def __init__(self, browser_settings: BrowserSettings, export_dir: str | Path | None = None):
        self.browser_settings = browser_settings
        self.export_dir = Path(export_dir) if export_dir else Path.cwd() / "brand_exports"
        self.browser = BrowserManager(browser_settings)
        self.extractor: ListingExtractor | None = None

    async def setup(self) -> None:
        page = await self.browser.setup()
        self.extractor = ListingExtractor(page)

    async def close(self) -> None:
        await self.browser.close()

    @staticmethod
    def _export_has_rows(path: Path) -> bool:
        if not path.exists():
            return False
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            return sheet.max_row > 1
        except Exception:
            return True

    async def reset_browser(self) -> None:
        await self.close()
        self.browser = BrowserManager(self.browser_settings)
        self.extractor = None

    async def parse_brand(
        self,
        brand: str,
        source_urls: list[str],
        pages: int,
        min_unique_cards: int,
        seen_articles: set[str],
    ) -> pd.DataFrame:
        if not self.extractor:
            await self.setup()

        out_path = self.export_dir / brand_file_name(brand)
        if out_path.exists():
            if self._export_has_rows(out_path):
                logger.warning("[%s] Файл %s уже существует, пропускаю бренд", brand, out_path.name)
                return pd.DataFrame()
            logger.warning("[%s] Файл %s пустой, перезаписываю", brand, out_path.name)

        writer = StreamingXlsxWriter(out_path)
        collected: dict[str, ListingItem] = {}

        def register(item: ListingItem) -> bool:
            article = article_from_url(item.url)
            if not article or article in seen_articles:
                return False
            seen_articles.add(article)
            collected[item.url] = item
            return True

        async def handle_new_items(items: list[ListingItem]) -> None:
            rows = []
            for item in items:
                if register(item):
                    rows.append(
                        normalize_listing_item(
                            item,
                            brand_hint=brand,
                        )
                    )
            if rows:
                writer.append_many(rows)
                logger.info("[%s] Записано +%s новых карточек", brand, len(rows))

        try:
            for base_url in source_urls:
                for source_url in listing_url_variants(base_url):
                    logger.info("[%s] Источник URL: %s", brand, source_url)
                    page_no_growth_streak = 0
                    for page_num in range(1, pages + 1):
                        before_count = len(collected)
                        page_url = build_page_url(source_url, page_num)
                        items = await self.extractor.fetch_listing_page(
                            page_url,
                            brand_hint=brand,
                            on_new_items=handle_new_items,
                        )
                        for item in items:
                            register(item)
                        logger.info("[%s] Всего уникальных по бренду: %s", brand, len(collected))
                        if len(collected) == before_count:
                            page_no_growth_streak += 1
                        else:
                            page_no_growth_streak = 0
                        if len(collected) >= min_unique_cards:
                            break
                        if page_no_growth_streak >= 3:
                            logger.info(
                                "[%s] 3 страницы подряд без новых карточек, перехожу дальше",
                                brand,
                            )
                            break
                        await self.extractor.human_delay(2, 4)
                    if len(collected) >= min_unique_cards:
                        break
                if len(collected) >= min_unique_cards:
                    break
        finally:
            writer.close()

        rows = [
            normalize_listing_item(
                item,
                brand_hint=brand,
            )
            for item in collected.values()
        ]
        return pd.DataFrame(rows)

    @staticmethod
    def build_jobs(
        urls_by_brand: dict[str, list[str]],
        brands: list[str] | None = None,
        url_override: str | None = None,
    ) -> list[ParseJob]:
        selected_brands = brands or list(BRAND_URLS.keys())
        jobs: list[ParseJob] = []
        for brand in selected_brands:
            source_urls = [url_override] if url_override else urls_by_brand.get(brand, [])
            if not source_urls:
                logger.info("[%s] Ссылка не задана, пропускаю", brand)
                continue
            jobs.append(ParseJob(brand=brand, source_urls=source_urls))
        return jobs

    async def parse_jobs(
        self,
        jobs: list[ParseJob],
        pages: int,
        min_cards: int,
        use_brand_min_cards: bool = True,
    ) -> dict[str, int]:
        seen_articles: set[str] = set()
        results: dict[str, int] = {}

        if not jobs:
            return results

        await self.setup()
        logger.info("Единый проход: задач к обходу %s", len(jobs))

        for job in jobs:
            required_cards = (
                max(min_cards, BRAND_MIN_CARDS.get(job.brand, DEFAULT_MIN_CARDS))
                if use_brand_min_cards
                else min_cards
            )
            try:
                frame = await self.parse_brand(
                    brand=job.brand,
                    source_urls=job.source_urls,
                    pages=pages,
                    min_unique_cards=required_cards,
                    seen_articles=seen_articles,
                )
            except Exception as exc:
                logger.exception("[%s] Ошибка парсинга бренда, продолжаю следующий", job.brand)
                results[job.brand] = -1
                if "Target" in str(exc) or "closed" in str(exc).lower():
                    await self.reset_browser()
                continue
            results[job.brand] = len(frame)
            if len(frame) < required_cards:
                logger.warning(
                    "[%s] Собрано %s из ожидаемых %s",
                    job.brand,
                    len(frame),
                    required_cards,
                )

        logger.info("Итог по брендам: %s", results)
        return results

    async def run_once(
        self,
        urls_by_brand: dict[str, list[str]],
        brands: list[str] | None = None,
        pages: int = 20,
        min_cards: int = 200,
        url_override: str | None = None,
        use_brand_min_cards: bool = True,
    ) -> tuple[Path | None, int]:
        self.export_dir.mkdir(parents=True, exist_ok=True)
        jobs = self.build_jobs(
            urls_by_brand=urls_by_brand,
            brands=brands,
            url_override=url_override,
        )

        try:
            await self.parse_jobs(
                jobs=jobs,
                pages=pages,
                min_cards=min_cards,
                use_brand_min_cards=use_brand_min_cards,
            )
        finally:
            await self.close()

        return aggregate_brand_exports(self.export_dir)

    async def run_daily(self, hour: int = 8, minute: int = 0, **kwargs) -> None:
        while True:
            run_at = next_run_at(hour=hour, minute=minute)
            wait_seconds = max(0.0, (run_at - datetime.now()).total_seconds())
            logger.info("Следующий запуск: %s", run_at.strftime("%Y-%m-%d %H:%M:%S"))
            await asyncio.sleep(wait_seconds)
            try:
                await self.run_once(**kwargs)
            except Exception:
                logger.exception("Ошибка при плановом запуске")
